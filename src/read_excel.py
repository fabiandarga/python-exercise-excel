from openpyxl import load_workbook

# 1. Excel-Datei öffnen
workbook = load_workbook('mitarbeiter.xlsx')

# 2. Aktives Arbeitsblatt auswählen
sheet = workbook.active

# 3. Alternativ: Bestimmtes Arbeitsblatt nach Namen auswählen
# sheet = workbook['Tabelle1']

# 4. Einzelne Zelle lesen
wert_a1 = sheet['A1'].value
# oder
wert_a1 = sheet.cell(row=1, column=1).value

# 5. Mehrere Zellen als Bereich lesen
zellen = sheet['A1:C5']

# 6. Zeile für Zeile lesen
for row in sheet.rows:
    for cell in row:
        print(cell.value)

# 7. Maximale Anzahl von Zeilen und Spalten ermitteln
max_rows = sheet.max_row
max_cols = sheet.max_column
